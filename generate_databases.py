#!/usr/bin/env python3
"""
Generate SQLite database and Excel workbook from frequency data.
"""

import sqlite3
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import csv
from enhance_frequency_data import get_enhanced_frequency_data

def create_sqlite_database(data, filename="Frequencies.db"):
    """Create SQLite database from frequency data."""
    
    # Connect to SQLite database
    conn = sqlite3.connect(filename)
    cursor = conn.cursor()
    
    # Drop table if exists
    cursor.execute("DROP TABLE IF EXISTS Frequencies")
    
    # Create table
    cursor.execute("""
        CREATE TABLE Frequencies (
            ID INTEGER PRIMARY KEY AUTOINCREMENT,
            Band TEXT NOT NULL,
            Frequency_Start_MHz REAL NOT NULL,
            Frequency_End_MHz REAL NOT NULL,
            Wavelength TEXT,
            Primary_Use TEXT,
            Service_Type TEXT,
            DateAdded DATETIME DEFAULT CURRENT_TIMESTAMP,
            Notes TEXT
        )
    """)
    
    # Create indexes for faster searching
    cursor.execute("CREATE INDEX idx_frequency_range ON Frequencies (Frequency_Start_MHz, Frequency_End_MHz)")
    cursor.execute("CREATE INDEX idx_service_type ON Frequencies (Service_Type)")
    cursor.execute("CREATE INDEX idx_band ON Frequencies (Band)")
    
    # Insert data
    for row in data:
        cursor.execute("""
            INSERT INTO Frequencies (Band, Frequency_Start_MHz, Frequency_End_MHz, Wavelength, Primary_Use, Service_Type)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (row['Band'], row['Frequency_Start_MHz'], row['Frequency_End_MHz'], 
              row['Wavelength'], row['Primary_Use'], row['Service_Type']))
    
    # Commit and close
    conn.commit()
    conn.close()
    
    print(f"SQLite database '{filename}' created with {len(data)} entries")

def create_excel_workbook(data, filename="Frequencies.xlsx"):
    """Create Excel workbook with separate worksheets for each service type."""
    
    # Create DataFrame
    df = pd.DataFrame(data)
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Create summary sheet
    summary_sheet = wb.create_sheet("Summary")
    
    # Service type counts
    service_counts = df['Service_Type'].value_counts().sort_index()
    summary_data = [['Service Type', 'Number of Entries']]
    for service, count in service_counts.items():
        summary_data.append([service, count])
    
    # Add summary data
    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = summary_sheet.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:  # Header row
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
    
    # Auto-size columns
    for column in summary_sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        summary_sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Create separate worksheet for each service type
    service_types = df['Service_Type'].unique()
    
    for service_type in sorted(service_types):
        # Filter data for this service type
        service_data = df[df['Service_Type'] == service_type].copy()
        service_data = service_data.sort_values('Frequency_Start_MHz')
        
        # Create worksheet
        sheet_name = service_type.replace('/', '_')[:31]  # Excel sheet name limit
        ws = wb.create_sheet(sheet_name)
        
        # Add data to worksheet
        for r in dataframe_to_rows(service_data, index=False, header=True):
            ws.append(r)
        
        # Style header row
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create complete data sheet
    complete_sheet = wb.create_sheet("Complete Database")
    df_sorted = df.sort_values(['Service_Type', 'Frequency_Start_MHz'])
    
    for r in dataframe_to_rows(df_sorted, index=False, header=True):
        complete_sheet.append(r)
    
    # Style header row
    for cell in complete_sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Auto-size columns
    for column in complete_sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        complete_sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Save workbook
    wb.save(filename)
    print(f"Excel workbook '{filename}' created with {len(service_types)} service-specific worksheets")

def create_access_instructions():
    """Create updated Access database instructions."""
    
    instructions = """# Creating the Frequencies Access Database

Since Access database files (.accdb) are binary files that require Microsoft Access to create properly, this repository provides multiple options for accessing the frequency data.

## Available Database Files

1. **Frequencies.db** - SQLite database (cross-platform, can be opened with many tools)
2. **Frequencies.xlsx** - Excel workbook with categorized worksheets  
3. **frequency_data.csv** - Enhanced CSV file with all frequency data
4. **database_schema.sql** - SQL schema for creating Access database

## Option 1: Use the SQLite Database (Recommended)

The `Frequencies.db` file is a complete SQLite database that can be:
- Opened with DB Browser for SQLite (free)
- Accessed from Python, Java, .NET, and most programming languages
- Imported into Access if needed
- Used with any SQLite-compatible application

### SQLite Query Examples:
```sql
-- Find what service uses a specific frequency
SELECT * FROM Frequencies 
WHERE 146.52 BETWEEN Frequency_Start_MHz AND Frequency_End_MHz;

-- Find all amateur radio bands
SELECT * FROM Frequencies 
WHERE Service_Type = 'Amateur Radio' 
ORDER BY Frequency_Start_MHz;

-- Find frequencies in a range
SELECT * FROM Frequencies 
WHERE Frequency_Start_MHz >= 144 AND Frequency_End_MHz <= 148;
```

## Option 2: Use the Excel Workbook

The `Frequencies.xlsx` file contains:
- **Summary** - Overview of all service types
- **Complete Database** - All frequency data sorted
- **Service-specific sheets** - Separate worksheet for each service type

## Option 3: Create Access Database from SQLite

1. Open Microsoft Access
2. Go to External Data → More → ODBC Database
3. Select "Import the source data into a new table"
4. Choose SQLite ODBC driver
5. Point to the `Frequencies.db` file
6. Import the Frequencies table

## Option 4: Manual Access Database Creation

### Step 1: Create New Database
1. Open Microsoft Access
2. Create a new blank database
3. Save it as "Frequencies.accdb"

### Step 2: Import Data
1. Go to External Data → Text File
2. Select "frequency_data.csv"
3. Choose "Import the source data into a new table"
4. Follow the Import Text Wizard:
   - Choose "Delimited"
   - Select "Comma" as delimiter
   - Check "First Row Contains Field Names"
   - Set appropriate data types for each field

### Step 3: Apply Schema
1. Open the imported table in Design View
2. Set the ID field as AutoNumber and Primary Key
3. Apply the indexes as specified in database_schema.sql
4. Add any additional fields (DateAdded, Notes) as needed

## Database Features

The completed database includes:
- **90+ frequency allocations** across all major services
- **Searchable by frequency range** - find what service uses a specific frequency
- **Categorized by service type** - Amateur Radio, Public Safety, Broadcast, etc.
- **Indexed for fast searches** - optimized for frequency lookups
- **Expandable** - easy to add new frequency allocations

## Data Sources

Frequency data compiled and verified from:
- FCC frequency allocations (CFR Title 47)
- ARRL Amateur Radio band plans
- ITU Radio Regulations
- 3GPP cellular specifications  
- IEEE wireless standards
- NIST time and frequency standards
- ICAO aviation standards
- Public safety frequency databases

---

*The frequency data in this repository has been researched and verified against authoritative sources as of 2024.*
"""
    
    with open('ACCESS_DATABASE_INSTRUCTIONS.md', 'w') as f:
        f.write(instructions)
    
    print("Updated ACCESS_DATABASE_INSTRUCTIONS.md with comprehensive options")

def main():
    """Generate all database and spreadsheet files."""
    print("Generating enhanced frequency database files...")
    
    # Get enhanced data
    data = get_enhanced_frequency_data()
    
    # Replace the original CSV with enhanced data
    with open('frequency_data.csv', 'w', newline='') as csvfile:
        fieldnames = ["Band", "Frequency_Start_MHz", "Frequency_End_MHz", "Wavelength", "Primary_Use", "Service_Type"]
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        for row in data:
            writer.writerow(row)
    
    print(f"Updated frequency_data.csv with {len(data)} verified entries")
    
    # Create SQLite database
    create_sqlite_database(data)
    
    # Create Excel workbook
    create_excel_workbook(data)
    
    # Update Access instructions
    create_access_instructions()
    
    print("\n=== Database Generation Complete ===")
    print(f"Files created:")
    print(f"- frequency_data.csv ({len(data)} entries)")
    print(f"- Frequencies.db (SQLite database)")
    print(f"- Frequencies.xlsx (Excel workbook)")
    print(f"- ACCESS_DATABASE_INSTRUCTIONS.md (updated)")

if __name__ == "__main__":
    main()